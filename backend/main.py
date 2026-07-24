from fastapi import FastAPI, HTTPException, Header, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, EmailStr
from datetime import datetime, timedelta, timezone
from pathlib import Path
import sqlite3
import hashlib
import jwt
import re
import os
import secrets
import time
import json
from urllib import error as urllib_error
from urllib import request as urllib_request
from dotenv import load_dotenv
from backend import rate_limit

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:
    psycopg = None
    dict_row = None

# =========================
# PATH
# =========================
BASE_DIR = Path(__file__).resolve().parent.parent
BACKEND_DIR = BASE_DIR / "backend"
FRONTEND_DIR = BASE_DIR / "frontend"
SQL_DIR = BASE_DIR / "sql"
DB_PATH = SQL_DIR / "app.db"
load_dotenv(BASE_DIR / ".env")


def csv_env(name: str):
    return [item.strip() for item in os.getenv(name, "").split(",") if item.strip()]


def bool_env(name: str, default: bool = False):
    raw_value = os.getenv(name)
    if raw_value is None:
        return default
    return raw_value.strip().lower() in {"1", "true", "yes", "on"}


DEFAULT_FRONTEND_ORIGINS = [
    "http://127.0.0.1:8000",
    "http://localhost:8000",
    "http://127.0.0.1:5500",
    "http://localhost:3000",
]
FRONTEND_ORIGINS = csv_env("FRONTEND_ORIGINS") or DEFAULT_FRONTEND_ORIGINS
FRONTEND_ORIGIN_REGEX = os.getenv("FRONTEND_ORIGIN_REGEX")

# =========================
# APP
# =========================
APP_VERSION = "5.0.0"

app = FastAPI(title="Card Education System API", version=APP_VERSION)

# Mount frontend assets under /static so API routes remain unambiguous.
# Frontend files are available through paths like /static/js/api.js.
app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")

# =========================
# CONFIG
# =========================
SECRET_KEY = os.getenv("SECRET_KEY", "change_me_please_use_env_in_production")
ALGORITHM = "HS256"
TOKEN_EXPIRE_DAYS = 3650
TAIPEI_TZ = timezone(timedelta(hours=8))

# =========================
# RATE LIMIT
# =========================
REDIS_URL = os.getenv("REDIS_URL", "").strip()
RATE_LIMIT_ENABLED = bool_env("RATE_LIMIT_ENABLED", True) and bool(REDIS_URL)
RATE_LIMIT_FAIL_OPEN = bool_env("RATE_LIMIT_FAIL_OPEN", False)
RATE_LIMIT_KEY_PREFIX = os.getenv("RATE_LIMIT_KEY_PREFIX", "cardlearn:rate-limit").strip()
if not RATE_LIMIT_KEY_PREFIX:
    RATE_LIMIT_KEY_PREFIX = "cardlearn:rate-limit"

rate_limit.configure_rate_limit(
    redis_url=REDIS_URL,
    enabled=RATE_LIMIT_ENABLED,
    fail_open=RATE_LIMIT_FAIL_OPEN,
    key_prefix=RATE_LIMIT_KEY_PREFIX,
)

GLOBAL_API_RATE_LIMIT = rate_limit.GLOBAL_API_RATE_LIMIT
LOGIN_RATE_LIMIT = rate_limit.LOGIN_RATE_LIMIT
REGISTER_RATE_LIMIT = rate_limit.REGISTER_RATE_LIMIT
SEND_OTP_RATE_LIMIT = rate_limit.SEND_OTP_RATE_LIMIT
GAME_RECORD_RATE_LIMIT = rate_limit.GAME_RECORD_RATE_LIMIT
API_RATE_LIMIT_PREFIXES = ("/auth", "/content", "/game", "/teacher", "/health")


def get_client_ip(request: Request):
    forwarded_for = request.headers.get("x-forwarded-for", "")
    if forwarded_for:
        return forwarded_for.split(",", 1)[0].strip()

    real_ip = request.headers.get("x-real-ip", "").strip()
    if real_ip:
        return real_ip

    if request.client:
        return request.client.host
    return "unknown"


def is_api_rate_limited_path(path: str):
    return any(path == prefix or path.startswith(f"{prefix}/") for prefix in API_RATE_LIMIT_PREFIXES)


def raise_rate_limit_unavailable():
    raise HTTPException(status_code=503, detail="Rate limit service unavailable.")


def enforce_rate_limit(rule: rate_limit.RateLimitRule, identifier: str):
    try:
        rate_limit.enforce_rate_limit(rule, identifier)
    except rate_limit.RateLimitBackendUnavailable:
        raise_rate_limit_unavailable()
    except rate_limit.RateLimitExceeded as error:
        raise HTTPException(
            status_code=429,
            detail="Too many requests. Please try again later.",
            headers=rate_limit.rate_limit_headers(error.result),
        )


@app.middleware("http")
async def global_api_rate_limit(request: Request, call_next):
    if request.method != "OPTIONS" and is_api_rate_limited_path(request.url.path):
        try:
            result = rate_limit.record_rate_limit_hit(
                GLOBAL_API_RATE_LIMIT,
                f"ip:{get_client_ip(request)}",
            )
        except rate_limit.RateLimitBackendUnavailable:
            return JSONResponse(
                status_code=503,
                content={"detail": "Rate limit service unavailable."},
            )

        if result is not None and not result.allowed:
            return JSONResponse(
                status_code=429,
                content={"detail": "Too many requests. Please try again later."},
                headers=rate_limit.rate_limit_headers(result),
            )

    return await call_next(request)


app.add_middleware(
    CORSMiddleware,
    allow_origins=FRONTEND_ORIGINS,
    allow_origin_regex=FRONTEND_ORIGIN_REGEX,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# EMAIL CONFIG
# =========================
# Gmail example:
# SMTP_HOST=smtp.gmail.com
# SMTP_PORT=587
# SMTP_USER=your Gmail
# SMTP_PASSWORD=your Google App Password, not your Gmail login password
# SMTP_FROM=CardLearn SDGs <your Gmail>
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "").strip()
RESEND_FROM = os.getenv("RESEND_FROM", "").strip() or os.getenv("SMTP_FROM", "").strip()

# =========================
# DB
# =========================
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)


def prepare_sql(sql: str):
    if not USE_POSTGRES:
        return sql
    return re.sub(r"\?", "%s", sql)


def auto_pk_sql():
    if USE_POSTGRES:
        return "INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY"
    return "INTEGER PRIMARY KEY AUTOINCREMENT"


class PostgresCursor:
    def __init__(self, cursor):
        self.cursor = cursor

    def execute(self, sql: str, params=()):
        self.cursor.execute(prepare_sql(sql), params)
        return self

    def executemany(self, sql: str, params):
        self.cursor.executemany(prepare_sql(sql), params)
        return self

    def fetchone(self):
        return self.cursor.fetchone()

    def fetchall(self):
        return self.cursor.fetchall()


class PostgresConnection:
    def __init__(self, conn):
        self.conn = conn

    def cursor(self):
        return PostgresCursor(self.conn.cursor())

    def commit(self):
        self.conn.commit()

    def close(self):
        self.conn.close()


def get_db():
    if USE_POSTGRES:
        if psycopg is None:
            raise RuntimeError("DATABASE_URL is set, but psycopg is not installed")
        return PostgresConnection(psycopg.connect(DATABASE_URL, row_factory=dict_row))

    SQL_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 30000")
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        conn.execute("PRAGMA journal_mode = WAL")
    except sqlite3.OperationalError:
        # Another uvicorn reload process may be opening the DB. The timeout
        # above still protects normal reads/writes; WAL can be enabled later.
        pass
    return conn


def execute_with_retry(cur, sql: str, params=(), retries: int = 6):
    if USE_POSTGRES:
        return cur.execute(sql, params)

    for attempt in range(retries):
        try:
            return cur.execute(sql, params)
        except sqlite3.OperationalError as error:
            if "database is locked" not in str(error).lower() or attempt == retries - 1:
                raise
            time.sleep(0.5 * (attempt + 1))


def table_columns(cur, table_name: str):
    if USE_POSTGRES:
        cur.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = ?
            """,
            (table_name,),
        )
        return [row["column_name"] for row in cur.fetchall()]

    cur.execute(f"PRAGMA table_info({table_name})")
    return [column["name"] for column in cur.fetchall()]


def execute_insert_returning(cur, sql: str, params, id_column: str):
    if USE_POSTGRES:
        cur.execute(f"{sql.strip()} RETURNING {id_column}", params)
        row = cur.fetchone()
        return row[id_column]

    cur.execute(sql, params)
    return cur.lastrowid


def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS users (
            id {auto_pk_sql()},
            name TEXT NOT NULL,
            gender TEXT,
            birth_date TEXT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            email_verified INTEGER DEFAULT 0,
            privacy_agreed INTEGER DEFAULT 0,
            role TEXT NOT NULL DEFAULT 'student',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS email_otps (
            id {auto_pk_sql()},
            email TEXT UNIQUE NOT NULL,
            otp TEXT NOT NULL,
            verified INTEGER DEFAULT 0,
            expires_at TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS sdg_options (
            sdg_level INTEGER PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT,
            active INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS difficulty_options (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            description TEXT,
            active INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0
        )
        """
    )

    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS tbl_sdg_parts (
            part_id {auto_pk_sql()},
            sdg_level INTEGER NOT NULL,
            part_no INTEGER NOT NULL,
            part_description TEXT NOT NULL,
            UNIQUE(sdg_level, part_no),
            FOREIGN KEY (sdg_level) REFERENCES sdg_options(sdg_level)
        )
        """
    )

    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS tbl_sdg_sub_topics (
            sub_topic_id {auto_pk_sql()},
            part_id INTEGER NOT NULL,
            sub_topic_no INTEGER NOT NULL,
            sub_topic_description TEXT NOT NULL,
            UNIQUE(part_id, sub_topic_no),
            FOREIGN KEY (part_id) REFERENCES tbl_sdg_parts(part_id)
        )
        """
    )

    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS tbl_sdg_cards (
            card_id {auto_pk_sql()},
            sub_topic_id INTEGER NOT NULL,
            difficulty_code TEXT NOT NULL,
            card_text TEXT NOT NULL,
            active INTEGER DEFAULT 1,
            UNIQUE(sub_topic_id, difficulty_code),
            FOREIGN KEY (sub_topic_id) REFERENCES tbl_sdg_sub_topics(sub_topic_id),
            FOREIGN KEY (difficulty_code) REFERENCES difficulty_options(code)
        )
        """
    )

    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS game_sessions (
            session_id {auto_pk_sql()},
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL,
            sdg_level INTEGER NOT NULL,
            difficulty_code TEXT NOT NULL,
            started_at TEXT NOT NULL,
            ended_at TEXT,
            status TEXT NOT NULL DEFAULT 'active',
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        """
    )

    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS game_part_records (
            id {auto_pk_sql()},
            session_id INTEGER NOT NULL,
            part_id INTEGER NOT NULL,
            errors INTEGER NOT NULL DEFAULT 0,
            repeat_listens INTEGER NOT NULL DEFAULT 0,
            duration INTEGER NOT NULL DEFAULT 0,
            completed_at TEXT NOT NULL,
            UNIQUE(session_id, part_id),
            FOREIGN KEY (session_id) REFERENCES game_sessions(session_id)
        )
        """
    )

    # Older app.db files may not have the role column yet.
    # Add it automatically for legacy tables.
    columns = table_columns(cur, "users")

    if "role" not in columns:
        execute_with_retry(cur, "ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'student'")

    game_part_columns = table_columns(cur, "game_part_records")
    if "repeat_listens" not in game_part_columns:
        execute_with_retry(
            cur,
            "ALTER TABLE game_part_records ADD COLUMN repeat_listens INTEGER NOT NULL DEFAULT 0",
        )

    game_session_columns = table_columns(cur, "game_sessions")
    if "user_id" not in game_session_columns:
        execute_with_retry(cur, "ALTER TABLE game_sessions ADD COLUMN user_id INTEGER")

    execute_with_retry(
        cur,
        """
        UPDATE game_sessions
        SET user_id = (
            SELECT MIN(users.id)
            FROM users
            WHERE users.name = game_sessions.username
        )
        WHERE user_id IS NULL
          AND 1 = (
            SELECT COUNT(*)
            FROM users
            WHERE users.name = game_sessions.username
          )
        """,
    )

    cur.execute("CREATE INDEX IF NOT EXISTS idx_game_sessions_user_id ON game_sessions(user_id)")

    seed_sdg_options = [
        (4, "Quality Education", "Quality Education", 1, 1),
        (13, "Climate Action", "Climate Action", 1, 2),
        (5, "Gender Equality", "Gender Equality", 1, 3),
    ]
    cur.executemany(
        """
        INSERT INTO sdg_options (sdg_level, title, description, active, sort_order)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(sdg_level) DO NOTHING
        """,
        seed_sdg_options,
    )
    cur.executemany(
        """
        UPDATE sdg_options
        SET description = ?
        WHERE sdg_level = ?
          AND (description IS NULL OR description = ?)
        """,
        [
            ("Quality Education", 4, "\u512a\u8cea\u6559\u80b2"),
            ("Climate Action", 13, "\u6c23\u5019\u884c\u52d5"),
            ("Gender Equality", 5, "\u6027\u5225\u5e73\u7b49"),
        ],
    )

    seed_difficulties = [
        ("A2", "A2", "A2 level", 1, 1),
        ("A1", "A1", "Beginner", 1, 2),
        ("B1", "B1", "Intermediate", 1, 3),
    ]
    cur.executemany(
        """
        INSERT INTO difficulty_options (code, label, description, active, sort_order)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(code) DO NOTHING
        """,
        seed_difficulties,
    )
    for code, label, description, active, sort_order in seed_difficulties:
        cur.execute(
            """
            UPDATE difficulty_options
            SET label = ?, description = ?, active = ?, sort_order = ?
            WHERE code = ?
            """,
            (label, description, active, sort_order, code),
        )
    cur.execute(
        """
        UPDATE difficulty_options
        SET active = 0, sort_order = 99
        WHERE code = ?
        """,
        ("A2-2",),
    )

    seed_parts = [
        (4, 1, "Introduction"),
        (4, 2, "Problems"),
        (4, 3, "Achievements"),
        (4, 4, "Solutions"),
        (4, 5, "Implications"),
        (4, 6, "Conclusion"),
        (13, 1, "Introduction"),
        (13, 2, "Problems"),
        (13, 3, "Achievements"),
        (13, 4, "Solutions"),
        (13, 5, "Implications"),
        (13, 6, "Conclusion"),
        (5, 1, "Introduction"),
        (5, 2, "Problems"),
        (5, 3, "Achievements"),
        (5, 4, "Solutions"),
        (5, 5, "Implications"),
        (5, 6, "Conclusion"),
    ]
    cur.executemany(
        """
        INSERT INTO tbl_sdg_parts (sdg_level, part_no, part_description)
        VALUES (?, ?, ?)
        ON CONFLICT(sdg_level, part_no) DO NOTHING
        """,
        seed_parts,
    )

    conn.commit()
    conn.close()


init_db()

# =========================
# UTIL
# =========================
def now_taipei():
    return datetime.now(TAIPEI_TZ)


def normalize_email(email: str):
    return email.strip().lower()


def generate_otp():
    return f"{secrets.randbelow(900000) + 100000:06d}"


def send_email(to_email: str, subject: str, body: str):
    if not SMTP_USER or not SMTP_PASSWORD:
        raise HTTPException(
            status_code=500,
            detail="SMTP is not configured. Please set SMTP_USER and SMTP_PASSWORD.",
        )

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = SMTP_FROM
    message["To"] = to_email
    message.set_content(body)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as smtp:
            smtp.starttls()
            smtp.login(SMTP_USER, SMTP_PASSWORD)
            smtp.send_message(message)
    except Exception as error:
        print("[EMAIL ERROR]", error)
        raise HTTPException(status_code=500, detail="Email delivery failed. Please try again later.")


def send_otp_email(to_email: str, otp: str):
    subject = "CardLearn SDGs Email Verification Code"
    body = f"""
Hello,

Your CardLearn SDGs email verification code is: {otp}

This code will expire in 5 minutes.
If you did not request this code, you can ignore this email.

CardLearn SDGs
""".strip()

    send_email(to_email, subject, body)


def send_email(to_email: str, subject: str, body: str):
    if not RESEND_API_KEY or not RESEND_FROM:
        raise HTTPException(
            status_code=500,
            detail="Email service is not configured. Please set RESEND_API_KEY and RESEND_FROM.",
        )

    payload = json.dumps(
        {
            "from": RESEND_FROM,
            "to": [to_email],
            "subject": subject,
            "text": body,
        }
    ).encode("utf-8")
    req = urllib_request.Request(
        "https://api.resend.com/emails",
        data=payload,
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
            "User-Agent": "CardLearn-SDGs/1.0",
        },
        method="POST",
    )

    try:
        with urllib_request.urlopen(req, timeout=20) as response:
            response_body = response.read().decode("utf-8")
            response_data = json.loads(response_body) if response_body else {}
            print(f"[EMAIL] provider=resend to={to_email} id={response_data.get('id', '')}")
    except urllib_error.HTTPError as err:
        error_body = err.read().decode("utf-8", errors="replace")
        print(f"[EMAIL ERROR] provider=resend status={err.code} body={error_body}")
        raise HTTPException(status_code=500, detail="Email delivery failed. Please try again later.")
    except Exception as err:
        print(f"[EMAIL ERROR] provider=resend error={err}")
        raise HTTPException(status_code=500, detail="Email delivery failed. Please try again later.")


def send_otp_email(to_email: str, otp: str):
    subject = "CardLearn SDGs Email Verification Code"
    body = f"""
Hello,

Your CardLearn SDGs verification code is: {otp}

This code will expire in 5 minutes.
If you did not request this code, you can ignore this email.

CardLearn SDGs
""".strip()
    send_email(to_email, subject, body)


def create_admin_user(
    name: str,
    email: str,
    password: str,
    gender: str | None = None,
    birth_date: str | None = None
):
    """
    Create an admin account directly from the backend.
    """
    """
    Create an admin account directly from the backend.
    Usage: temporarily call create_admin_user(...) once near the bottom
    of the app, then remove or comment that call after the account is created.
    """
    email = normalize_email(email)

    if not validate_password(password):
        raise ValueError("Admin password must be at least 8 characters and include letters and numbers.")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT id FROM users WHERE email = ?", (email,))
    existing_user = cur.fetchone()

    if existing_user:
        conn.close()
        print(f"[ADMIN] {email} already exists; skipping duplicate creation.")
        return

    cur.execute(
        """
        INSERT INTO users (
            name,
            gender,
            birth_date,
            email,
            password_hash,
            email_verified,
            privacy_agreed,
            role
        ) VALUES (?, ?, ?, ?, ?, 1, 1, 'admin')
        """,
        (
            name.strip(),
            gender,
            birth_date,
            email,
            hash_password(password),
        ),
    )

    conn.commit()
    conn.close()

    print(f"[ADMIN] Admin account created successfully: {email}")


# =========================
# SECURITY
# =========================
def hash_password(password: str):
    return hashlib.sha256(password.encode()).hexdigest()


def verify_password(password: str, hashed: str):
    return hash_password(password) == hashed


def validate_password(password: str):
    return (
        len(password) >= 8
        and re.search(r"[A-Za-z]", password)
        and re.search(r"\d", password)
    )


def create_jwt(user_id: int):
    payload = {
        "user_id": user_id,
        "exp": datetime.utcnow() + timedelta(days=TOKEN_EXPIRE_DAYS),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def decode_jwt(token: str):
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalid")


def get_token_from_authorization(authorization: str | None):
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")

    parts = authorization.split()
    if len(parts) != 2 or parts[0].lower() != "bearer":
        raise HTTPException(status_code=401, detail="Invalid Authorization format")

    return parts[1]


def get_current_user(authorization: str | None):
    token = get_token_from_authorization(authorization)
    payload = decode_jwt(token)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name, email, role FROM users WHERE id = ?", (payload["user_id"],))
    user = cur.fetchone()
    conn.close()

    if not user:
        raise HTTPException(status_code=404, detail="user not found")

    return dict(user)


def require_teacher(authorization: str | None):
    user = get_current_user(authorization)
    if user.get("role") not in ("teacher", "admin"):
        raise HTTPException(status_code=403, detail="teacher permission required")
    return user


def can_access_game_session(session, current_user: dict):
    if current_user.get("role") in ("teacher", "admin"):
        return True
    return session["user_id"] is not None and session["user_id"] == current_user["id"]

# =========================
# SCHEMAS
# =========================
class SendOTPIn(BaseModel):
    email: EmailStr


class VerifyOTPIn(BaseModel):
    email: EmailStr
    otp: str


class RegisterIn(BaseModel):
    name: str
    gender: str | None = None
    birth_date: str | None = None
    email: EmailStr
    password: str
    confirm_password: str
    privacy_agreed: bool


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class GameStartIn(BaseModel):
    username: str
    sdg: int
    level: str


class GamePartIn(BaseModel):
    session_id: int
    part_id: int
    errors: int = 0
    repeat_listens: int = 0
    duration: int = 0


class GameEndIn(BaseModel):
    session_id: int


class SdgOptionIn(BaseModel):
    sdg_level: int
    title: str
    description: str | None = None
    active: bool = True


class DifficultyOptionIn(BaseModel):
    code: str
    label: str
    description: str | None = None
    active: bool = True
    sort_order: int = 0


class SdgPartIn(BaseModel):
    sdg_level: int
    part_no: int
    part_description: str


class SdgSubTopicIn(BaseModel):
    part_id: int
    sub_topic_no: int
    sub_topic_description: str


class SdgCardIn(BaseModel):
    sub_topic_id: int
    difficulty_code: str
    card_text: str
    active: bool = True


class UserRoleIn(BaseModel):
    role: str

# =========================
# PAGES
# =========================
@app.get("/")
def home():
    return FileResponse(FRONTEND_DIR / "index.html")


@app.get("/pages/register.html")
def register_page():
    return FileResponse(FRONTEND_DIR / "pages" / "register.html")


@app.get("/pages/login.html")
def login_page():
    return FileResponse(FRONTEND_DIR / "pages" / "login.html")


@app.get("/pages/{page_name}")
def frontend_page(page_name: str):
    page_path = FRONTEND_DIR / "pages" / page_name
    if not page_path.is_file() or page_path.suffix.lower() != ".html":
        raise HTTPException(status_code=404, detail="page not found")
    return FileResponse(page_path)

# =========================
# ROUTES
# =========================
@app.get("/health")
def health_check():
    return {
        "message": "API is running",
        "database": "postgresql" if USE_POSTGRES else str(DB_PATH),
        "rate_limit": "redis" if RATE_LIMIT_ENABLED else "disabled",
        "time": now_taipei().isoformat(),
    }

# -------------------------
# SEND OTP
# -------------------------
@app.post("/auth/send-otp")
def send_otp(data: SendOTPIn):
    email = normalize_email(data.email)
    enforce_rate_limit(SEND_OTP_RATE_LIMIT, f"email:{email}")

    otp = generate_otp()
    expires_at = now_taipei() + timedelta(minutes=5)

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO email_otps (email, otp, verified, expires_at)
        VALUES (?, ?, 0, ?)
        ON CONFLICT(email) DO UPDATE SET
            otp = excluded.otp,
            verified = 0,
            expires_at = excluded.expires_at,
            created_at = CURRENT_TIMESTAMP
        """,
        (email, otp, expires_at.isoformat()),
    )

    conn.commit()
    conn.close()

    print(f"[OTP] email={email}, expires_at={expires_at.isoformat()}")

    send_otp_email(email, otp)

    return {
        "message": "OTP sent. Please check your inbox.",
        "expires_in_minutes": 5,
    }

# -------------------------
# VERIFY OTP
# -------------------------
@app.post("/auth/verify-otp")
def verify_otp(data: VerifyOTPIn):
    email = normalize_email(data.email)
    otp = data.otp.strip()

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM email_otps WHERE email = ?", (email,))
    record = cur.fetchone()

    if not record:
        conn.close()
        raise HTTPException(status_code=400, detail="Please request a verification code first.")

    expires_at = datetime.fromisoformat(record["expires_at"])
    if now_taipei() > expires_at:
        conn.close()
        raise HTTPException(status_code=400, detail="The verification code has expired. Please request a new one.")

    if otp != record["otp"]:
        conn.close()
        raise HTTPException(status_code=400, detail="The verification code is incorrect.")

    cur.execute("UPDATE email_otps SET verified = 1 WHERE email = ?", (email,))

    conn.commit()
    conn.close()

    return {"message": "Email verified successfully."}

# -------------------------
# REGISTER
# -------------------------
@app.post("/auth/register")
def register(data: RegisterIn, request: Request):
    email = normalize_email(data.email)
    enforce_rate_limit(REGISTER_RATE_LIMIT, f"ip:{get_client_ip(request)}")

    if not data.name.strip():
        raise HTTPException(status_code=400, detail="Please enter your name.")

    if not validate_password(data.password):
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters and include letters and numbers.")

    if data.password != data.confirm_password:
        raise HTTPException(status_code=400, detail="Passwords do not match.")

    if not data.privacy_agreed:
        raise HTTPException(status_code=400, detail="You must agree to the Privacy Policy.")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT id FROM users WHERE email = ?", (email,))
    existing_user = cur.fetchone()
    if existing_user:
        conn.close()
        raise HTTPException(status_code=400, detail="Email already exists.")

    cur.execute("SELECT * FROM email_otps WHERE email = ?", (email,))
    otp_record = cur.fetchone()
    if not otp_record or otp_record["verified"] != 1:
        conn.close()
        raise HTTPException(status_code=400, detail="Please complete email verification first.")

    try:
        cur.execute(
            """
            INSERT INTO users (
                name,
                gender,
                birth_date,
                email,
                password_hash,
                email_verified,
                privacy_agreed,
                role
            ) VALUES (?, ?, ?, ?, ?, 1, ?, 'student')
            """,
            (
                data.name.strip(),
                data.gender,
                data.birth_date,
                email,
                hash_password(data.password),
                1,
            ),
        )

        conn.commit()

    except Exception as error:
        is_integrity_error = isinstance(error, sqlite3.IntegrityError) or (
            psycopg is not None and isinstance(error, psycopg.IntegrityError)
        )
        if not is_integrity_error:
            conn.close()
            raise
        conn.close()
        raise HTTPException(status_code=400, detail="Email already exists.")

    conn.close()

    return {"message": "Registration successful."}

# -------------------------
# LOGIN
# -------------------------
@app.post("/auth/login")
def login(data: LoginIn, request: Request):
    email = normalize_email(data.email)
    enforce_rate_limit(LOGIN_RATE_LIMIT, f"ip:{get_client_ip(request)}")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM users WHERE email = ?", (email,))
    user = cur.fetchone()

    conn.close()

    if not user:
        raise HTTPException(status_code=400, detail="Account does not exist.")

    if not verify_password(data.password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="Incorrect password.")

    token = create_jwt(user["id"])

    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user["id"],
            "name": user["name"],
            "email": user["email"],
            "username": user["name"],
            "role": user["role"],
        },
    }

# -------------------------
# ME
# -------------------------
@app.get("/auth/me")
def me(authorization: str | None = Header(default=None)):
    token = get_token_from_authorization(authorization)
    payload = decode_jwt(token)

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "SELECT id, name, email, gender, birth_date, email_verified, role, created_at FROM users WHERE id = ?",
        (payload["user_id"],),
    )
    user = cur.fetchone()

    conn.close()

    if not user:
        raise HTTPException(status_code=404, detail="user not found")

    return dict(user)


# -------------------------
# CONTENT OPTIONS
# -------------------------
@app.get("/content/options")
def content_options():
    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT sdg_level, title, description, active
        FROM sdg_options
        WHERE active = 1
        ORDER BY sdg_level
        """
    )
    sdgs = [dict(row) for row in cur.fetchall()]

    cur.execute(
        """
        SELECT code, label, description, active, sort_order
        FROM difficulty_options
        WHERE active = 1
        ORDER BY sort_order, code
        """
    )
    difficulties = [dict(row) for row in cur.fetchall()]

    conn.close()
    return {"sdgs": sdgs, "difficulties": difficulties}


@app.get("/content/parts")
def content_parts(sdg: int | None = None):
    conn = get_db()
    cur = conn.cursor()

    params = []
    where = ""
    if sdg is not None:
        where = "WHERE p.sdg_level = ?"
        params.append(sdg)

    cur.execute(
        f"""
        SELECT
            p.part_id,
            p.sdg_level,
            p.part_no,
            p.part_description,
            st.sub_topic_id,
            st.sub_topic_no,
            st.sub_topic_description,
            c.card_id,
            c.difficulty_code,
            c.card_text,
            c.active AS card_active
        FROM tbl_sdg_parts p
        LEFT JOIN tbl_sdg_sub_topics st ON st.part_id = p.part_id
        LEFT JOIN tbl_sdg_cards c ON c.sub_topic_id = st.sub_topic_id
        {where}
        ORDER BY p.sdg_level, p.part_no, st.sub_topic_no, c.difficulty_code
        """,
        params,
    )
    rows = cur.fetchall()
    conn.close()

    parts_by_id = {}
    for row in rows:
        part_id = row["part_id"]
        if part_id not in parts_by_id:
            parts_by_id[part_id] = {
                "part_id": part_id,
                "sdg_level": row["sdg_level"],
                "part_no": row["part_no"],
                "part_description": row["part_description"],
                "sub_topics": [],
            }
        if row["sub_topic_id"] is not None:
            sub_topic = next(
                (
                    item for item in parts_by_id[part_id]["sub_topics"]
                    if item["sub_topic_id"] == row["sub_topic_id"]
                ),
                None,
            )
            if sub_topic is None:
                sub_topic = {
                    "sub_topic_id": row["sub_topic_id"],
                    "sub_topic_no": row["sub_topic_no"],
                    "sub_topic_description": row["sub_topic_description"],
                    "cards": [],
                }
                parts_by_id[part_id]["sub_topics"].append(sub_topic)

            if row["card_id"] is not None:
                sub_topic["cards"].append({
                    "card_id": row["card_id"],
                    "difficulty_code": row["difficulty_code"],
                    "card_text": row["card_text"],
                    "active": bool(row["card_active"]),
                })

    return {"parts": list(parts_by_id.values())}


@app.get("/game/content")
def game_content(sdg: int, level: str, authorization: str | None = Header(default=None)):
    get_current_user(authorization)

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT sdg_level, title, description
        FROM sdg_options
        WHERE sdg_level = ? AND active = 1
        """,
        (sdg,),
    )
    sdg_row = cur.fetchone()
    if not sdg_row:
        conn.close()
        raise HTTPException(status_code=404, detail="SDG not found")

    cur.execute(
        """
        SELECT code, label, description
        FROM difficulty_options
        WHERE code = ? AND active = 1
        """,
        (level,),
    )
    difficulty_row = cur.fetchone()
    if not difficulty_row:
        conn.close()
        raise HTTPException(status_code=404, detail="difficulty not found")

    cur.execute(
        """
        SELECT
            p.part_id,
            p.part_no,
            p.part_description,
            st.sub_topic_id,
            st.sub_topic_no,
            st.sub_topic_description,
            c.card_id,
            c.card_text
        FROM tbl_sdg_parts p
        LEFT JOIN tbl_sdg_sub_topics st ON st.part_id = p.part_id
        LEFT JOIN tbl_sdg_cards c
            ON c.sub_topic_id = st.sub_topic_id
            AND c.difficulty_code = ?
            AND c.active = 1
        WHERE p.sdg_level = ?
        ORDER BY p.part_no, st.sub_topic_no
        """,
        (level, sdg),
    )
    rows = cur.fetchall()
    conn.close()

    parts_by_id = {}
    for row in rows:
        part_id = row["part_id"]
        if part_id not in parts_by_id:
            parts_by_id[part_id] = {
                "part_id": part_id,
                "part_no": row["part_no"],
                "part_description": row["part_description"],
                "sub_topics": [],
            }
        if row["sub_topic_id"] is not None:
            parts_by_id[part_id]["sub_topics"].append({
                "sub_topic_id": row["sub_topic_id"],
                "sub_topic_no": row["sub_topic_no"],
                "sub_topic_description": row["sub_topic_description"],
                "card": {
                    "card_id": row["card_id"],
                    "card_text": row["card_text"],
                    "difficulty_code": level,
                } if row["card_id"] is not None else None,
            })

    return {
        "sdg": dict(sdg_row),
        "difficulty": dict(difficulty_row),
        "parts": list(parts_by_id.values()),
    }


# -------------------------
# GAME
# -------------------------
@app.post("/game/start")
def start_game(data: GameStartIn, authorization: str | None = Header(default=None)):
    current_user = get_current_user(authorization)

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT sdg_level FROM sdg_options WHERE sdg_level = ? AND active = 1", (data.sdg,))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="SDG option not found")

    cur.execute("SELECT code FROM difficulty_options WHERE code = ? AND active = 1", (data.level,))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="difficulty option not found")

    started_at = now_taipei().isoformat()
    session_id = execute_insert_returning(
        cur,
        """
        INSERT INTO game_sessions (user_id, username, sdg_level, difficulty_code, started_at, status)
        VALUES (?, ?, ?, ?, ?, 'active')
        """,
        (current_user["id"], current_user["name"], data.sdg, data.level, started_at),
        "session_id",
    )
    conn.commit()
    conn.close()

    return {
        "session_id": session_id,
        "sdg": data.sdg,
        "level": data.level,
        "started_at": started_at,
    }


@app.post("/game/part")
def save_game_part(data: GamePartIn, authorization: str | None = Header(default=None)):
    current_user = get_current_user(authorization)
    enforce_rate_limit(GAME_RECORD_RATE_LIMIT, f"user:{current_user['id']}")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT session_id, user_id, username FROM game_sessions WHERE session_id = ?", (data.session_id,))
    session = cur.fetchone()
    if not session:
        conn.close()
        raise HTTPException(status_code=404, detail="session not found")
    if not can_access_game_session(session, current_user):
        conn.close()
        raise HTTPException(status_code=403, detail="session permission required")

    cur.execute(
        """
        INSERT INTO game_part_records (session_id, part_id, errors, repeat_listens, duration, completed_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(session_id, part_id) DO UPDATE SET
            errors = excluded.errors,
            repeat_listens = excluded.repeat_listens,
            duration = excluded.duration,
            completed_at = excluded.completed_at
        """,
        (
            data.session_id,
            data.part_id,
            max(0, data.errors),
            max(0, data.repeat_listens),
            max(0, data.duration),
            now_taipei().isoformat(),
        ),
    )

    conn.commit()
    conn.close()
    return {"message": "part saved"}


@app.post("/game/end")
def end_game(data: GameEndIn, authorization: str | None = Header(default=None)):
    current_user = get_current_user(authorization)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT session_id, user_id, username FROM game_sessions WHERE session_id = ?", (data.session_id,))
    session = cur.fetchone()
    if not session:
        conn.close()
        raise HTTPException(status_code=404, detail="session not found")
    if not can_access_game_session(session, current_user):
        conn.close()
        raise HTTPException(status_code=403, detail="session permission required")

    cur.execute(
        """
        UPDATE game_sessions
        SET status = 'completed', ended_at = ?
        WHERE session_id = ?
        """,
        (now_taipei().isoformat(), data.session_id),
    )
    conn.commit()
    conn.close()
    return {"message": "game ended"}


@app.get("/game/results/{session_id}")
def game_results(session_id: int, authorization: str | None = Header(default=None)):
    current_user = get_current_user(authorization)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM game_sessions WHERE session_id = ?", (session_id,))
    session = cur.fetchone()
    if not session:
        conn.close()
        raise HTTPException(status_code=404, detail="session not found")
    if not can_access_game_session(session, current_user):
        conn.close()
        raise HTTPException(status_code=403, detail="result permission required")

    cur.execute(
        """
        SELECT part_id, errors, repeat_listens, duration, completed_at
        FROM game_part_records
        WHERE session_id = ?
        ORDER BY part_id
        """,
        (session_id,),
    )
    parts = [dict(row) for row in cur.fetchall()]
    cur.execute(
        """
        SELECT
            r.part_id,
            gs.username,
            gs.session_id,
            r.errors,
            r.repeat_listens,
            r.duration,
            ROW_NUMBER() OVER (
                PARTITION BY r.part_id
                ORDER BY r.errors ASC, r.duration ASC, r.repeat_listens ASC, gs.started_at ASC
            ) AS rank
        FROM game_part_records r
        JOIN game_sessions gs ON gs.session_id = r.session_id
        WHERE gs.sdg_level = ? AND gs.difficulty_code = ? AND gs.status = 'completed'
        ORDER BY r.part_id, rank
        """,
        (session["sdg_level"], session["difficulty_code"]),
    )
    leaderboard_rows = [dict(row) for row in cur.fetchall()]
    conn.close()

    leaderboard = {}
    for row in leaderboard_rows:
        part_key = str(row["part_id"])
        leaderboard.setdefault(part_key, [])
        leaderboard[part_key].append({
            "rank": row["rank"],
            "display_name": "You" if row["session_id"] == session_id else f"Student {row['rank']}",
            "is_current_user": row["session_id"] == session_id,
            "errors": row["errors"],
            "repeat_listens": row["repeat_listens"],
            "duration": row["duration"],
        })

    return {
        "session": dict(session),
        "parts": parts,
        "total_errors": sum(part["errors"] for part in parts),
        "total_repeat_listens": sum(part["repeat_listens"] for part in parts),
        "total_duration": sum(part["duration"] for part in parts),
        "leaderboard": leaderboard,
    }


# -------------------------
# TEACHER CONTENT MANAGEMENT
# -------------------------
@app.get("/teacher/content")
def teacher_content(authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    return {
        **content_options(),
        **content_parts(),
    }


def clean_excel_value(value):
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def normalize_excel_difficulty(value):
    text = clean_excel_value(value)
    if not text:
        return ""
    match = re.search(r"\b(A1|A2|B1)\b", text, re.IGNORECASE)
    if not match:
        return ""
    code = match.group(1).upper()
    return code


def parse_part_number(value):
    text = clean_excel_value(value)
    if not text:
        return None
    match = re.search(r"part\s*(\d+)", text, re.IGNORECASE)
    if match:
        return int(match.group(1))
    if text.isdigit():
        return int(text)
    return None


def parse_sdg_sheet_title(title, first_row):
    title_text = clean_excel_value(title)
    first_cells = [clean_excel_value(cell) for cell in first_row if clean_excel_value(cell)]
    first_text = first_cells[0] if first_cells else title_text
    match = re.search(r"SDG\s*(\d+)", f"{title_text} {first_text}", re.IGNORECASE)
    if not match:
        return None, "", ""

    sdg_level = int(match.group(1))
    label_source = first_text if "SDG" in first_text.upper() else title_text
    if ":" in label_source:
        title_part = label_source.split(":", 1)[1].strip()
    else:
        title_part = re.sub(r"SDG\s*\d+", "", label_source, flags=re.IGNORECASE).strip(" :-")
    title_part = title_part or f"SDG {sdg_level}"
    return sdg_level, title_part, title_part


def row_value(row, index):
    if index < 0 or index >= len(row):
        return ""
    return clean_excel_value(row[index])


def find_part_column(rows):
    for row in rows[1:20]:
        for index, value in enumerate(row):
            if parse_part_number(value):
                return index
    return 0


def xlsx_column_index(cell_ref):
    match = re.match(r"([A-Z]+)", cell_ref or "")
    if not match:
        return 0
    index = 0
    for char in match.group(1):
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def load_xlsx_rows_stdlib(file_bytes: bytes):
    import io
    import zipfile
    import xml.etree.ElementTree as ET

    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    ns_pkg_rel = "{http://schemas.openxmlformats.org/package/2006/relationships}"

    def read_xml(archive, path):
        return ET.fromstring(archive.read(path))

    with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = read_xml(archive, "xl/sharedStrings.xml")
            for item in root.findall(f"{ns_main}si"):
                text_parts = [node.text or "" for node in item.iter(f"{ns_main}t")]
                shared_strings.append("".join(text_parts))

        workbook_root = read_xml(archive, "xl/workbook.xml")
        rels_root = read_xml(archive, "xl/_rels/workbook.xml.rels")
        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root.findall(f"{ns_pkg_rel}Relationship")
        }

        sheets = []
        for sheet in workbook_root.findall(f".//{ns_main}sheet"):
            name = sheet.attrib.get("name", "")
            rel_id = sheet.attrib.get(f"{ns_rel}id")
            target = rel_targets.get(rel_id, "")
            if not target:
                continue
            if target.startswith("/"):
                sheet_path = target.lstrip("/")
            else:
                sheet_path = f"xl/{target}"
            sheets.append((name, sheet_path))

        workbook_rows = []
        for name, sheet_path in sheets:
            root = read_xml(archive, sheet_path)
            rows = []
            for row_node in root.findall(f".//{ns_main}sheetData/{ns_main}row"):
                row_values = []
                for cell in row_node.findall(f"{ns_main}c"):
                    col_index = xlsx_column_index(cell.attrib.get("r", ""))
                    while len(row_values) <= col_index:
                        row_values.append(None)

                    cell_type = cell.attrib.get("t")
                    value = ""
                    if cell_type == "inlineStr":
                        value = "".join(node.text or "" for node in cell.iter(f"{ns_main}t"))
                    else:
                        value_node = cell.find(f"{ns_main}v")
                        value = value_node.text if value_node is not None else ""
                        if cell_type == "s" and value != "":
                            value = shared_strings[int(value)]

                    row_values[col_index] = value
                rows.append(row_values)
            workbook_rows.append((name, rows))

    return workbook_rows


def load_question_bank_rows(file_bytes: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return load_xlsx_rows_stdlib(file_bytes)

    try:
        import io
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Unable to read Excel workbook") from exc

    return [
        (sheet.title, list(sheet.iter_rows(values_only=True)))
        for sheet in workbook.worksheets
    ]


def parse_question_bank_workbook(file_bytes: bytes):
    workbook_rows = load_question_bank_rows(file_bytes)

    imported = {
        "sdgs": {},
        "difficulties": set(),
        "parts": [],
        "sub_topics": [],
        "cards": [],
    }

    for sheet_title, rows in workbook_rows:
        if not rows:
            continue

        first_row = rows[0]
        sdg_level, sdg_title, sdg_description = parse_sdg_sheet_title(sheet_title, first_row)
        if not sdg_level:
            continue

        difficulty_columns = {}
        for index, value in enumerate(first_row):
            difficulty = normalize_excel_difficulty(value)
            if difficulty:
                difficulty_columns[index] = difficulty
                imported["difficulties"].add(difficulty)

        if not difficulty_columns:
            continue

        first_difficulty_col = min(difficulty_columns)
        part_col = find_part_column(rows)
        part_description_col = part_col + 1
        sub_topic_no_col = part_col + 2
        sub_topic_title_start_col = part_col + 3
        imported["sdgs"][sdg_level] = {
            "sdg_level": sdg_level,
            "title": sdg_title,
            "description": sdg_description,
        }

        current_part_no = None
        current_part_description = ""

        for row in rows[1:]:
            part_no = parse_part_number(row_value(row, part_col))
            part_description = row_value(row, part_description_col)
            if part_no:
                current_part_no = part_no
                current_part_description = part_description or current_part_description
                imported["parts"].append({
                    "sdg_level": sdg_level,
                    "part_no": current_part_no,
                    "part_description": current_part_description or f"Part {current_part_no}",
                })

            if not current_part_no:
                continue

            sub_topic_raw = row_value(row, sub_topic_no_col)
            if not sub_topic_raw:
                continue
            try:
                sub_topic_no = int(float(sub_topic_raw))
            except ValueError:
                continue

            sub_topic_description = ""
            for index in range(sub_topic_title_start_col, first_difficulty_col):
                candidate = row_value(row, index)
                if candidate:
                    sub_topic_description = candidate
                    break
            if not sub_topic_description:
                sub_topic_description = f"Subtopic {sub_topic_no}"

            imported["sub_topics"].append({
                "sdg_level": sdg_level,
                "part_no": current_part_no,
                "sub_topic_no": sub_topic_no,
                "sub_topic_description": sub_topic_description,
            })

            for column_index, difficulty_code in difficulty_columns.items():
                card_text = row_value(row, column_index)
                if not card_text:
                    continue
                imported["cards"].append({
                    "sdg_level": sdg_level,
                    "part_no": current_part_no,
                    "sub_topic_no": sub_topic_no,
                    "difficulty_code": difficulty_code,
                    "card_text": card_text,
                })

    if not imported["cards"]:
        raise HTTPException(status_code=400, detail="No importable sentence cards were found in this workbook")

    return imported


def first_material_value(*values):
    for value in values:
        text = clean_excel_value(value)
        if text:
            return text
    return ""


def parse_material_sdg_level(value):
    text = clean_excel_value(value)
    if not text:
        return None
    match = re.search(r"\d+", text)
    if not match:
        return None
    return int(match.group(0))


def parse_material_sub_topic_no(value, part_no, fallback):
    text = clean_excel_value(value)
    if not text:
        return fallback
    if "." in text:
        text = text.rsplit(".", 1)[-1]
    try:
        return int(float(text))
    except ValueError:
        match = re.search(r"\d+", text)
        if match:
            return int(match.group(0))
    return fallback


def parse_material_json(file_bytes: bytes):
    try:
        payload = json.loads(file_bytes.decode("utf-8-sig"))
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail="JSON file must be UTF-8 encoded") from exc
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: line {exc.lineno}, column {exc.colno}") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Material JSON root must be an object")

    sdg = payload.get("sdg") if isinstance(payload.get("sdg"), dict) else {}
    difficulty = payload.get("difficulty") if isinstance(payload.get("difficulty"), dict) else {}
    if not difficulty and isinstance(payload.get("level"), dict):
        difficulty = payload.get("level")

    sdg_level = parse_material_sdg_level(
        sdg.get("level")
        or sdg.get("number")
        or sdg.get("code")
        or payload.get("sdg_level")
        or payload.get("sdg_code")
    )
    if not sdg_level:
        raise HTTPException(status_code=400, detail="Material JSON must include sdg.level or sdg.code")

    difficulty_code = normalize_excel_difficulty(
        difficulty.get("code")
        or payload.get("difficulty_code")
        or payload.get("level_code")
    )
    if not difficulty_code:
        raise HTTPException(status_code=400, detail="Material JSON difficulty must be A1, A2, or B1")

    parts = payload.get("parts")
    if not isinstance(parts, list) or not parts:
        raise HTTPException(status_code=400, detail="Material JSON must include a non-empty parts array")

    imported = {
        "sdgs": {
            sdg_level: {
                "sdg_level": sdg_level,
                "title": first_material_value(sdg.get("title"), sdg.get("name"), f"SDG {sdg_level}"),
                "description": first_material_value(sdg.get("description"), sdg.get("zh_name"), sdg.get("title"), f"SDG {sdg_level}"),
            }
        },
        "difficulties": {difficulty_code},
        "parts": [],
        "sub_topics": [],
        "cards": [],
    }
    errors = []
    seen_parts = set()

    for part_index, part in enumerate(parts, start=1):
        if not isinstance(part, dict):
            errors.append(f"Part item {part_index}: must be an object")
            continue

        part_no = parse_part_number(part.get("part_no") or part.get("no") or part.get("title") or part_index)
        if not part_no or part_no < 1 or part_no > 6:
            errors.append(f"Part item {part_index}: part_no must be between 1 and 6")
            continue
        if part_no in seen_parts:
            errors.append(f"Part {part_no}: duplicated part_no")
            continue
        seen_parts.add(part_no)

        part_description = first_material_value(
            part.get("description"),
            part.get("title"),
            part.get("zh_title"),
            f"Part {part_no}",
        )
        imported["parts"].append({
            "sdg_level": sdg_level,
            "part_no": part_no,
            "part_description": part_description,
        })

        sub_topics = (
            part.get("sub_topics")
            or part.get("subtopics")
            or part.get("sub_parts")
            or part.get("subparts")
        )
        if not isinstance(sub_topics, list) or not sub_topics:
            errors.append(f"Part {part_no}: sub_topics must be a non-empty array")
            continue

        seen_sub_topics = set()
        for sub_index, sub_topic in enumerate(sub_topics, start=1):
            if not isinstance(sub_topic, dict):
                errors.append(f"Part {part_no}, subtopic {sub_index}: must be an object")
                continue

            sub_topic_no = parse_material_sub_topic_no(
                sub_topic.get("sub_topic_no")
                or sub_topic.get("subpart_no")
                or sub_topic.get("no"),
                part_no,
                sub_index,
            )
            if sub_topic_no in seen_sub_topics:
                errors.append(f"Part {part_no}, subtopic {sub_topic_no}: duplicated subtopic number")
                continue
            seen_sub_topics.add(sub_topic_no)

            card_text = first_material_value(
                sub_topic.get("card_text"),
                sub_topic.get("description"),
                sub_topic.get("text"),
                sub_topic.get("sentence"),
            )
            if not card_text:
                errors.append(f"Part {part_no}, subtopic {sub_topic_no}: card_text or description is required")
                continue

            sub_topic_description = first_material_value(
                sub_topic.get("title"),
                sub_topic.get("sub_topic_description"),
                sub_topic.get("subpart_title"),
                sub_topic.get("name"),
                f"Subtopic {sub_topic_no}",
            )
            imported["sub_topics"].append({
                "sdg_level": sdg_level,
                "part_no": part_no,
                "sub_topic_no": sub_topic_no,
                "sub_topic_description": sub_topic_description,
            })
            imported["cards"].append({
                "sdg_level": sdg_level,
                "part_no": part_no,
                "sub_topic_no": sub_topic_no,
                "difficulty_code": difficulty_code,
                "card_text": card_text,
            })

    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors[:8]))
    if not imported["cards"]:
        raise HTTPException(status_code=400, detail="No importable sentence cards were found in this JSON file")

    return imported


def difficulty_label_for_import(code):
    return code


def import_question_bank(parsed, source_label="Excel"):
    conn = get_db()
    cur = conn.cursor()
    summary = {
        "sdgs": len(parsed["sdgs"]),
        "difficulties": len(parsed["difficulties"]),
        "parts": 0,
        "sub_topics": 0,
        "cards": 0,
    }
    part_ids = {}
    sub_topic_ids = {}

    try:
        for sdg in parsed["sdgs"].values():
            cur.execute(
                """
                INSERT INTO sdg_options (sdg_level, title, description, active, sort_order)
                VALUES (?, ?, ?, 1, ?)
                ON CONFLICT(sdg_level) DO UPDATE SET
                    title = excluded.title,
                    description = excluded.description,
                    active = 1
                """,
                (sdg["sdg_level"], sdg["title"], sdg["description"], sdg["sdg_level"]),
            )

        for index, code in enumerate(sorted(parsed["difficulties"]), start=1):
            cur.execute(
                """
                INSERT INTO difficulty_options (code, label, description, active, sort_order)
                VALUES (?, ?, ?, 1, ?)
                ON CONFLICT(code) DO UPDATE SET
                    label = excluded.label,
                    description = excluded.description,
                    active = 1,
                    sort_order = excluded.sort_order
                """,
                (code, difficulty_label_for_import(code), f"{code} imported from {source_label}", index),
            )

        unique_parts = {
            (item["sdg_level"], item["part_no"]): item
            for item in parsed["parts"]
        }
        for key, part in unique_parts.items():
            cur.execute(
                """
                INSERT INTO tbl_sdg_parts (sdg_level, part_no, part_description)
                VALUES (?, ?, ?)
                ON CONFLICT(sdg_level, part_no) DO UPDATE SET
                    part_description = excluded.part_description
                """,
                (part["sdg_level"], part["part_no"], part["part_description"]),
            )
            cur.execute(
                "SELECT part_id FROM tbl_sdg_parts WHERE sdg_level = ? AND part_no = ?",
                (part["sdg_level"], part["part_no"]),
            )
            part_ids[key] = cur.fetchone()["part_id"]
        summary["parts"] = len(unique_parts)

        unique_sub_topics = {
            (item["sdg_level"], item["part_no"], item["sub_topic_no"]): item
            for item in parsed["sub_topics"]
        }
        for key, sub_topic in unique_sub_topics.items():
            part_id = part_ids[(sub_topic["sdg_level"], sub_topic["part_no"])]
            cur.execute(
                """
                INSERT INTO tbl_sdg_sub_topics (part_id, sub_topic_no, sub_topic_description)
                VALUES (?, ?, ?)
                ON CONFLICT(part_id, sub_topic_no) DO UPDATE SET
                    sub_topic_description = excluded.sub_topic_description
                """,
                (part_id, sub_topic["sub_topic_no"], sub_topic["sub_topic_description"]),
            )
            cur.execute(
                "SELECT sub_topic_id FROM tbl_sdg_sub_topics WHERE part_id = ? AND sub_topic_no = ?",
                (part_id, sub_topic["sub_topic_no"]),
            )
            sub_topic_ids[key] = cur.fetchone()["sub_topic_id"]
        summary["sub_topics"] = len(unique_sub_topics)

        unique_cards = {
            (item["sdg_level"], item["part_no"], item["sub_topic_no"], item["difficulty_code"]): item
            for item in parsed["cards"]
        }
        for key, card in unique_cards.items():
            sub_topic_id = sub_topic_ids[(card["sdg_level"], card["part_no"], card["sub_topic_no"])]
            cur.execute(
                """
                INSERT INTO tbl_sdg_cards (sub_topic_id, difficulty_code, card_text, active)
                VALUES (?, ?, ?, 1)
                ON CONFLICT(sub_topic_id, difficulty_code) DO UPDATE SET
                    card_text = excluded.card_text,
                    active = 1
                """,
                (sub_topic_id, card["difficulty_code"], card["card_text"]),
            )
        summary["cards"] = len(unique_cards)

        conn.commit()
    except Exception:
        conn.close()
        raise

    conn.close()
    return summary


@app.post("/teacher/import-excel")
async def import_teacher_excel(file: UploadFile = File(...), authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if len(content) > 12 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Uploaded file is too large")

    parsed = parse_question_bank_workbook(content)
    summary = import_question_bank(parsed, source_label="Excel")
    return {
        "message": "question bank imported",
        "filename": filename,
        "summary": summary,
    }


@app.post("/teacher/import-json")
async def import_teacher_json(file: UploadFile = File(...), authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    filename = file.filename or ""
    if not filename.lower().endswith(".json"):
        raise HTTPException(status_code=400, detail="Please upload a .json file")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Uploaded file is too large")

    parsed = parse_material_json(content)
    summary = import_question_bank(parsed, source_label="JSON")
    return {
        "message": "material JSON imported",
        "filename": filename,
        "summary": summary,
    }


@app.get("/teacher/game-records")
def teacher_game_records(authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            gs.session_id,
            gs.user_id,
            gs.username,
            gs.sdg_level,
            gs.difficulty_code,
            gs.started_at,
            gs.ended_at,
            gs.status,
            ROW_NUMBER() OVER (
                PARTITION BY COALESCE(CAST(gs.user_id AS TEXT), 'legacy:' || gs.username), gs.sdg_level, gs.difficulty_code
                ORDER BY gs.started_at ASC, gs.session_id ASC
            ) AS attempt_no,
            COALESCE(SUM(r.errors), 0) AS total_errors,
            COALESCE(SUM(r.repeat_listens), 0) AS total_repeat_listens,
            COALESCE(SUM(r.duration), 0) AS total_duration,
            COUNT(r.id) AS completed_parts
        FROM game_sessions gs
        LEFT JOIN game_part_records r ON r.session_id = gs.session_id
        WHERE gs.status = 'completed'
        GROUP BY gs.session_id
        ORDER BY gs.started_at DESC
        """
    )
    sessions = [dict(row) for row in cur.fetchall()]
    session_map = {item["session_id"]: item for item in sessions}
    for item in sessions:
        item["parts"] = []

    if session_map:
        placeholders = ",".join("?" for _ in session_map)
        cur.execute(
            f"""
            SELECT
                r.session_id,
                r.part_id,
                p.part_no,
                p.part_description,
                r.errors,
                r.repeat_listens,
                r.duration,
                r.completed_at,
                COUNT(c.card_id) AS card_count
            FROM game_part_records r
            JOIN game_sessions gs ON gs.session_id = r.session_id
            LEFT JOIN tbl_sdg_parts p ON p.part_id = r.part_id
            LEFT JOIN tbl_sdg_sub_topics st ON st.part_id = p.part_id
            LEFT JOIN tbl_sdg_cards c
                ON c.sub_topic_id = st.sub_topic_id
                AND c.difficulty_code = gs.difficulty_code
                AND c.active = 1
            WHERE r.session_id IN ({placeholders})
            GROUP BY
                r.session_id,
                r.part_id,
                p.part_no,
                p.part_description,
                r.errors,
                r.repeat_listens,
                r.duration,
                r.completed_at
            ORDER BY r.session_id DESC, COALESCE(p.part_no, r.part_id)
            """,
            tuple(session_map.keys()),
        )
        for row in cur.fetchall():
            item = dict(row)
            item["suspected_guess"] = item["card_count"] > 0 and item["errors"] >= item["card_count"] + 1
            session_map[row["session_id"]]["parts"].append(item)

    suspected_guess_parts = sum(
        1
        for session in sessions
        for part in session["parts"]
        if part.get("suspected_guess")
    )
    for session in sessions:
        session["suspected_guess_parts"] = sum(1 for part in session["parts"] if part.get("suspected_guess"))

    cur.execute(
        """
        SELECT
            gs.sdg_level,
            gs.difficulty_code,
            r.part_id,
            p.part_no,
            COUNT(*) AS attempts,
            ROUND(AVG(r.errors), 2) AS avg_errors,
            ROUND(AVG(r.repeat_listens), 2) AS avg_repeat_listens,
            ROUND(AVG(r.duration), 2) AS avg_duration
        FROM game_part_records r
        JOIN game_sessions gs ON gs.session_id = r.session_id
        LEFT JOIN tbl_sdg_parts p ON p.part_id = r.part_id
        WHERE gs.status = 'completed'
        GROUP BY gs.sdg_level, gs.difficulty_code, r.part_id, p.part_no
        ORDER BY gs.sdg_level, gs.difficulty_code, COALESCE(p.part_no, r.part_id)
        """
    )
    part_stats = [dict(row) for row in cur.fetchall()]
    conn.close()

    return {
        "sessions": sessions,
        "part_stats": part_stats,
        "overview": {
            "total_sessions": len(sessions),
            "completed_sessions": sum(1 for item in sessions if item["status"] == "completed"),
            "total_students": len({
                item["user_id"] if item.get("user_id") is not None else f"legacy:{item['username']}"
                for item in sessions
            }),
            "suspected_guess_parts": suspected_guess_parts,
        },
    }


@app.delete("/teacher/game-records/{session_id}")
def delete_teacher_game_record(session_id: int, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT session_id FROM game_sessions WHERE session_id = ?", (session_id,))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="session not found")
    cur.execute("DELETE FROM game_part_records WHERE session_id = ?", (session_id,))
    cur.execute("DELETE FROM game_sessions WHERE session_id = ?", (session_id,))
    conn.commit()
    conn.close()
    return {"message": "game record deleted"}


@app.get("/teacher/users")
def teacher_users(authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            u.id,
            u.name,
            u.gender,
            u.birth_date,
            u.email,
            u.email_verified,
            u.privacy_agreed,
            u.role,
            u.created_at,
            COUNT(gs.session_id) AS total_sessions,
            COALESCE(SUM(CASE WHEN gs.status = 'completed' THEN 1 ELSE 0 END), 0) AS completed_sessions,
            MAX(gs.started_at) AS last_game_at
        FROM users u
        LEFT JOIN game_sessions gs ON gs.user_id = u.id
        GROUP BY u.id
        ORDER BY u.created_at DESC, u.id DESC
        """
    )
    users = [dict(row) for row in cur.fetchall()]
    conn.close()
    return {
        "users": users,
        "overview": {
            "total_users": len(users),
            "students": sum(1 for user in users if user["role"] == "student"),
            "teachers": sum(1 for user in users if user["role"] == "teacher"),
            "admins": sum(1 for user in users if user["role"] == "admin"),
        },
    }


@app.patch("/teacher/users/{user_id}/role")
def update_teacher_user_role(user_id: int, data: UserRoleIn, authorization: str | None = Header(default=None)):
    current_user = require_teacher(authorization)
    new_role = data.role.strip().lower()
    if new_role not in ("student", "teacher", "admin"):
        raise HTTPException(status_code=400, detail="invalid role")
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="cannot change your own role")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, email, role FROM users WHERE id = ?", (user_id,))
    target_user = cur.fetchone()
    if not target_user:
        conn.close()
        raise HTTPException(status_code=404, detail="user not found")
    if (target_user["role"] == "admin" or new_role == "admin") and current_user.get("role") != "admin":
        conn.close()
        raise HTTPException(status_code=403, detail="admin role can only be managed by admin")

    cur.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
    conn.commit()
    conn.close()
    return {"message": "user role updated"}


@app.delete("/teacher/users/{user_id}")
def delete_teacher_user(user_id: int, authorization: str | None = Header(default=None)):
    current_user = require_teacher(authorization)
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="cannot delete your own account")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, email, role FROM users WHERE id = ?", (user_id,))
    target_user = cur.fetchone()
    if not target_user:
        conn.close()
        raise HTTPException(status_code=404, detail="user not found")
    if target_user["role"] == "admin" and current_user.get("role") != "admin":
        conn.close()
        raise HTTPException(status_code=403, detail="admin account can only be deleted by admin")

    cur.execute(
        """
        DELETE FROM game_part_records
        WHERE session_id IN (
            SELECT session_id
            FROM game_sessions
            WHERE user_id = ?
        )
        """,
        (user_id,),
    )
    cur.execute("DELETE FROM game_sessions WHERE user_id = ?", (user_id,))
    cur.execute("DELETE FROM email_otps WHERE email = ?", (target_user["email"],))
    cur.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return {"message": "user deleted"}


@app.post("/teacher/sdgs")
def upsert_sdg_option(data: SdgOptionIn, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO sdg_options (sdg_level, title, description, active, sort_order)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(sdg_level) DO UPDATE SET
            title = excluded.title,
            description = excluded.description,
            active = excluded.active
        """,
        (data.sdg_level, data.title.strip(), data.description, int(data.active), data.sdg_level),
    )
    conn.commit()
    conn.close()
    return {"message": "sdg saved"}


@app.post("/teacher/difficulties")
def upsert_difficulty_option(data: DifficultyOptionIn, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO difficulty_options (code, label, description, active, sort_order)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(code) DO UPDATE SET
            label = excluded.label,
            description = excluded.description,
            active = excluded.active,
            sort_order = excluded.sort_order
        """,
        (data.code.strip(), data.label.strip(), data.description, int(data.active), data.sort_order),
    )
    conn.commit()
    conn.close()
    return {"message": "difficulty saved"}


@app.post("/teacher/parts")
def upsert_sdg_part(data: SdgPartIn, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    if data.part_no < 1 or data.part_no > 6:
        raise HTTPException(status_code=400, detail="Part number must be between 1 and 6")
    conn = get_db()
    cur = conn.cursor()
    part_id = execute_insert_returning(
        cur,
        """
        INSERT INTO tbl_sdg_parts (sdg_level, part_no, part_description)
        VALUES (?, ?, ?)
        ON CONFLICT(sdg_level, part_no) DO UPDATE SET
            part_description = excluded.part_description
        """,
        (data.sdg_level, data.part_no, data.part_description.strip()),
        "part_id",
    )
    conn.commit()
    if not part_id:
        cur.execute(
            "SELECT part_id FROM tbl_sdg_parts WHERE sdg_level = ? AND part_no = ?",
            (data.sdg_level, data.part_no),
        )
        part_id = cur.fetchone()["part_id"]
    conn.close()
    return {"message": "part saved", "part_id": part_id}


@app.post("/teacher/sub-topics")
def upsert_sdg_sub_topic(data: SdgSubTopicIn, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    sub_topic_id = execute_insert_returning(
        cur,
        """
        INSERT INTO tbl_sdg_sub_topics (part_id, sub_topic_no, sub_topic_description)
        VALUES (?, ?, ?)
        ON CONFLICT(part_id, sub_topic_no) DO UPDATE SET
            sub_topic_description = excluded.sub_topic_description
        """,
        (data.part_id, data.sub_topic_no, data.sub_topic_description.strip()),
        "sub_topic_id",
    )
    conn.commit()
    if not sub_topic_id:
        cur.execute(
            "SELECT sub_topic_id FROM tbl_sdg_sub_topics WHERE part_id = ? AND sub_topic_no = ?",
            (data.part_id, data.sub_topic_no),
        )
        sub_topic_id = cur.fetchone()["sub_topic_id"]
    conn.close()
    return {"message": "sub topic saved", "sub_topic_id": sub_topic_id}


@app.post("/teacher/cards")
def upsert_sdg_card(data: SdgCardIn, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO tbl_sdg_cards (sub_topic_id, difficulty_code, card_text, active)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(sub_topic_id, difficulty_code) DO UPDATE SET
            card_text = excluded.card_text,
            active = excluded.active
        """,
        (
            data.sub_topic_id,
            data.difficulty_code.strip(),
            data.card_text.strip(),
            int(data.active),
        ),
    )
    conn.commit()
    conn.close()
    return {"message": "card saved"}


@app.delete("/teacher/sdgs/{sdg_level}")
def delete_sdg_option(sdg_level: int, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        DELETE FROM tbl_sdg_cards
        WHERE sub_topic_id IN (
            SELECT st.sub_topic_id
            FROM tbl_sdg_sub_topics st
            JOIN tbl_sdg_parts p ON p.part_id = st.part_id
            WHERE p.sdg_level = ?
        )
        """,
        (sdg_level,),
    )
    cur.execute(
        """
        DELETE FROM tbl_sdg_sub_topics
        WHERE part_id IN (SELECT part_id FROM tbl_sdg_parts WHERE sdg_level = ?)
        """,
        (sdg_level,),
    )
    cur.execute("DELETE FROM tbl_sdg_parts WHERE sdg_level = ?", (sdg_level,))
    cur.execute("DELETE FROM sdg_options WHERE sdg_level = ?", (sdg_level,))
    conn.commit()
    conn.close()
    return {"message": "sdg deleted"}


@app.delete("/teacher/parts/{part_id}")
def delete_sdg_part(part_id: int, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        DELETE FROM tbl_sdg_cards
        WHERE sub_topic_id IN (
            SELECT sub_topic_id FROM tbl_sdg_sub_topics WHERE part_id = ?
        )
        """,
        (part_id,),
    )
    cur.execute("DELETE FROM tbl_sdg_sub_topics WHERE part_id = ?", (part_id,))
    cur.execute("DELETE FROM tbl_sdg_parts WHERE part_id = ?", (part_id,))
    conn.commit()
    conn.close()
    return {"message": "part deleted"}


@app.delete("/teacher/sub-topics/{sub_topic_id}")
def delete_sdg_sub_topic(sub_topic_id: int, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM tbl_sdg_cards WHERE sub_topic_id = ?", (sub_topic_id,))
    cur.execute("DELETE FROM tbl_sdg_sub_topics WHERE sub_topic_id = ?", (sub_topic_id,))
    conn.commit()
    conn.close()
    return {"message": "sub topic deleted"}


@app.delete("/teacher/cards/{card_id}")
def delete_sdg_card(card_id: int, authorization: str | None = Header(default=None)):
    require_teacher(authorization)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM tbl_sdg_cards WHERE card_id = ?", (card_id,))
    conn.commit()
    conn.close()
    return {"message": "card deleted"}


# =========================
# CREATE INITIAL ADMIN
# =========================
INITIAL_ADMIN_EMAIL = os.getenv("INITIAL_ADMIN_EMAIL", "").strip()
INITIAL_ADMIN_PASSWORD = os.getenv("INITIAL_ADMIN_PASSWORD", "")
INITIAL_ADMIN_NAME = os.getenv("INITIAL_ADMIN_NAME", "Admin")

if INITIAL_ADMIN_EMAIL and INITIAL_ADMIN_PASSWORD:
    create_admin_user(
        name=INITIAL_ADMIN_NAME,
        email=INITIAL_ADMIN_EMAIL,
        password=INITIAL_ADMIN_PASSWORD,
    )
